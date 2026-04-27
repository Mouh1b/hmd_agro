"""
Parse "Rapport Mensuel" Excel workbooks and persist each daily sheet as a
`Rapport Journalier Importe` row. Read-only historical copy — the report
checks for imported data first, then falls back to live event queries.
"""

import io
import json
from datetime import date as date_cls

import frappe
import openpyxl

from hmd_agro.hmd_agro.utils.live_state import CATEGORIES, empty_row as _zero_row

DOCTYPE = "Rapport Journalier Importe"

# ─── Sheet layout ────────────────────────────────────────────────────────────
# Columns C..J carry the 8 category totals, in CATEGORIES order.
DATA_COLS = list(range(3, 11))  # C..J
assert len(DATA_COLS) == len(CATEGORIES)

# Map col A / col B labels (normalized) → key in the stored JSON.
# Labels come from the daily sheets; accents are stripped and case lowered.
ROW_LABELS = {
    ("effectif initial", ""):              "effectif_initial",
    ("changement de categorie", "+"):      "changement_cat_plus",
    ("", "-"):                             "changement_cat_minus",
    ("velage", ""):                        "velage",
    ("naissance", ""):                     "naissance",
    ("avortement / mort ne", ""):          "avortement_mort_ne",
    ("vente", "quantite"):                 "vente_qty",
    ("", "prix"):                          "vente_prix",
    ("motalite", ""):                      "mortalite",
    ("mortalite", ""):                     "mortalite",
    ("effectif final", ""):                "effectif_final",
}

# Keys that we expect to be set; anything missing is filled with zeros.
ALL_KEYS = [
    "effectif_initial", "changement_cat_plus", "changement_cat_minus",
    "velage", "naissance", "avortement_mort_ne",
    "achat", "vente_qty", "vente_prix", "mortalite", "reforme", "effectif_final",
]


# ─── Public API ──────────────────────────────────────────────────────────────

@frappe.whitelist()
def import_workbook(file_url, annee, mois):
    """
    Parse `file_url` and write one row per daily sheet named "01".."31".
    The date is built from (annee, mois, sheet_name) — day-cells inside
    sheets are ignored because the source Excel sometimes has date typos.
    """
    annee, mois = int(annee), int(mois)
    content = _load_file(file_url)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    imported, skipped = [], []
    for sheet_name in wb.sheetnames:
        day = _day_from_sheet_name(sheet_name)
        if not day:
            skipped.append({"sheet": sheet_name, "reason": "nom de feuille non numérique"})
            continue
        try:
            sheet_date = date_cls(annee, mois, day)
            ws = wb[sheet_name]
            data = _parse_effectif_block(ws)
            data["production_lot"] = _parse_production_lot_block(ws, sheet_date)
        except ValueError as e:
            skipped.append({"sheet": sheet_name, "reason": str(e)})
            continue

        _save(sheet_date, data, file_url)
        imported.append(str(sheet_date))

    frappe.db.commit()
    return {"imported": imported, "skipped": skipped}


def _day_from_sheet_name(name):
    try:
        d = int(name)
        return d if 1 <= d <= 31 else None
    except ValueError:
        return None


@frappe.whitelist()
def detect_period(file_url):
    """
    Read date cells in each daily sheet and return the most common (year, month).
    Used by the UI to pre-fill the import dialog.
    """
    from collections import Counter
    content = _load_file(file_url)
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    pairs = []
    for sheet_name in wb.sheetnames:
        if not _day_from_sheet_name(sheet_name):
            continue
        ws = wb[sheet_name]
        for row in range(1, 7):
            v = ws.cell(row=row, column=1).value
            if hasattr(v, "year") and hasattr(v, "month"):
                pairs.append((v.year, v.month))
                break
    if not pairs:
        return None
    annee, mois = Counter(pairs).most_common(1)[0][0]
    return {"annee": annee, "mois": mois, "confidence": pairs.count((annee, mois)), "total_sheets": len(pairs)}


def read_imported(date):
    """Return the imported rapport dict for `date`, or None."""
    raw = frappe.db.get_value(DOCTYPE, {"date": str(date)}, "rapport_json")
    return json.loads(raw) if raw else None


# ─── Parsing ─────────────────────────────────────────────────────────────────

def _parse_effectif_block(ws):
    """
    Walk rows in the sheet, match col A / col B labels to our keys, and
    extract the 8 category values from columns C..J.
    """
    data = {k: _zero_row() for k in ALL_KEYS}
    seen = set()

    for row in range(1, ws.max_row + 1):
        key = _match_row(
            _norm(ws.cell(row=row, column=1).value),
            _norm(ws.cell(row=row, column=2).value),
        )
        if not key or key in seen:
            continue
        seen.add(key)
        data[key] = _read_row_values(ws, row)

    if "effectif_initial" not in seen or "effectif_final" not in seen:
        raise ValueError("block Effectif incomplet (Initial ou Final manquant)")

    return data


def _match_row(label_a, label_b):
    return ROW_LABELS.get((label_a, label_b))


def _norm(v):
    """Normalize a cell value for label matching (None → '', strip accents/case)."""
    if v is None:
        return ""
    s = str(v).strip().lower()
    # Strip common accents without pulling in unicodedata
    for a, b in (("é","e"), ("è","e"), ("ê","e"), ("ë","e"),
                 ("à","a"), ("â","a"), ("î","i"), ("ï","i"),
                 ("ô","o"), ("ö","o"), ("û","u"), ("ù","u"), ("ç","c")):
        s = s.replace(a, b)
    return s


def _read_row_values(ws, row):
    """Read columns C..I from `row` into a CATEGORIES-shaped dict. Total is recomputed
    from the sum (Excel's Total cell is unreliable for change rows — often blank)."""
    out = {cat: 0 for cat in CATEGORIES}
    for cat, col in zip(CATEGORIES[:-1], DATA_COLS[:-1]):  # skip Total
        v = ws.cell(row=row, column=col).value
        try:
            out[cat] = int(round(float(v))) if v is not None else 0
        except (TypeError, ValueError):
            out[cat] = 0
    out["Total"] = sum(out[c] for c in CATEGORIES if c != "Total")
    return out


# ─── Production par Lot block ───────────────────────────────────────────────

_LOT_DATA_COLS = list(range(3, 10))  # C..I (lot columns, max 7 lots)


def _parse_production_lot_block(ws, sheet_date):
    """
    Extract per-lot effectif + production for `sheet_date`.
    Returns {lot_name: {"effectif": int, "production": float}} or {} if block missing.
    """
    lot_row = _find_row(ws, "lot")
    if not lot_row:
        return {}

    lots = []
    for col in _LOT_DATA_COLS:
        v = ws.cell(row=lot_row, column=col).value
        if v:
            lots.append((str(v).strip(), col))
    if not lots:
        return {}

    eff_row = _find_row(ws, "effectif", start=lot_row + 1, max_distance=2)
    prod_row = _find_date_row(ws, sheet_date, start=lot_row + 1, max_distance=5)

    result = {}
    for name, col in lots:
        eff = _num(ws.cell(row=eff_row, column=col).value) if eff_row else 0
        prod = _num(ws.cell(row=prod_row, column=col).value) if prod_row else 0
        result[name] = {"effectif": int(eff), "production": float(prod)}
    return result


def _find_row(ws, label, start=1, max_distance=None):
    end = (start + max_distance) if max_distance else ws.max_row + 1
    for row in range(start, min(end, ws.max_row + 1)):
        if _norm(ws.cell(row=row, column=1).value) == label:
            return row
    return None


def _find_date_row(ws, target_date, start=1, max_distance=None):
    end = (start + max_distance) if max_distance else ws.max_row + 1
    for row in range(start, min(end, ws.max_row + 1)):
        v = ws.cell(row=row, column=1).value
        if hasattr(v, "date") and v.date() == target_date:
            return row
        if isinstance(v, date_cls) and v == target_date:
            return row
    return None


def _num(v):
    if v is None:
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


# ─── Persistence ─────────────────────────────────────────────────────────────

def _save(date, data, source_file):
    date = str(date)
    payload = {"rapport_json": json.dumps(data), "source_file": source_file}
    existing = frappe.db.get_value(DOCTYPE, {"date": date})
    if existing:
        doc = frappe.get_doc(DOCTYPE, existing)
        doc.update(payload)
        doc.save(ignore_permissions=True)
    else:
        frappe.get_doc({"doctype": DOCTYPE, "date": date, **payload}).insert(ignore_permissions=True)


def _load_file(file_url):
    """Fetch file contents from Frappe File or accept an absolute path."""
    if file_url.startswith("/files/") or file_url.startswith("/private/"):
        f = frappe.get_doc("File", {"file_url": file_url})
        return f.get_content()
    with open(file_url, "rb") as fh:
        return fh.read()
